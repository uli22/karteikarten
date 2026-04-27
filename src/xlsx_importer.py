"""xlsx_importer.py — Importlogik für XLSX-Dateien (Taufbuch-Karteikarten)."""

from __future__ import annotations

import re
import unicodedata
from typing import Callable, Optional


def normalize_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def normalize_number(value) -> Optional[str]:
    if value is None:
        return None
    try:
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, int):
            return str(value)
    except Exception:
        pass
    return str(value).strip() if str(value).strip() else None


def normalize_year(value) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def normalize_date(value, wb_epoch) -> Optional[str]:
    """Normalisiert einen Datumswert auf DD.MM.YYYY.
    XX (unbekannter Tag/Monat) wird zu 00 konvertiert.
    Echte Excel-Datetime-Objekte werden direkt formatiert.
    Daten vor 1900 liegen als Strings vor (Excel-Limitierung)."""
    if value is None or str(value).strip() == "":
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    text = str(value).strip()
    text = re.sub(r'\bXX\b', '00', text, flags=re.IGNORECASE)
    match = re.match(r"^(\d{1,2})[\./-](\d{1,2})[\./-](\d{4})$", text)
    if match:
        day, month, year = match.groups()
        return f"{day.zfill(2)}.{month.zfill(2)}.{year}"
    match = re.match(r"^(\d{4})[\./-](\d{1,2})[\./-](\d{1,2})$", text)
    if match:
        year, month, day = match.groups()
        return f"{day.zfill(2)}.{month.zfill(2)}.{year}"
    return text


def iso_from_datum(datum) -> Optional[str]:
    """Wandelt DD.MM.YYYY in ISO-Format YYYY-MM-DD um. Unterstützt auch 00 für unbekannte Teile."""
    if not datum:
        return None
    datum = re.sub(r'\bXX\b', '00', datum, flags=re.IGNORECASE)
    match = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", datum)
    if match:
        day, month, year = match.groups()
        return f"{year}-{month}-{day}"
    match = re.match(r"^(\d{4})\.(\d{2})\.(\d{2})$", datum)
    if match:
        year, month, day = match.groups()
        return f"{year}-{month}-{day}"
    match = re.match(r"^(\d{4})$", datum)
    if match:
        return f"{match.group(1)}-00-00"
    return None


def to_ymd_dot(date_str) -> Optional[str]:
    """Wandelt DD.MM.YYYY in YYYY.MM.DD um (Format für todestag/datum_geburt).
    Unvollständige Daten (00.MM.YYYY) und reine Jahreszahlen bleiben erhalten."""
    if not date_str:
        return None
    match = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", date_str)
    if match:
        day, month, year = match.groups()
        return f"{year}.{month}.{day}"
    return date_str


def stand_from_gender(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"m", "maennlich", "männlich", "male", "1"}:
        return "Sohn"
    if text in {"w", "weiblich", "female", "f", "2"}:
        return "Tochter"
    if "m" in text and "weib" not in text:
        return "Sohn"
    if "w" in text or "weib" in text:
        return "Tochter"
    return None


def normalize_key(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\u00A0", " ")
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", " ", text)
    return text.casefold()


def build_match_keys(value) -> set:
    """Erzeugt alle möglichen Normalisierungsvarianten eines Dateinamens für den Abgleich."""
    keys = set()
    if value is None:
        return keys
    raw = str(value).strip()
    if not raw:
        return keys
    raw = re.sub(r"_erf(?=\.(jpg|jpeg|png|tif|tiff)$)", "", raw, flags=re.IGNORECASE)
    base = re.sub(r"\.(jpg|jpeg|png|tif|tiff)$", "", raw, flags=re.IGNORECASE)
    base_no_inf = re.sub(r"_inf$", "", base, flags=re.IGNORECASE)
    base_no_erf = re.sub(r"_erf$", "", base, flags=re.IGNORECASE)
    base_no_inf_erf = re.sub(r"_erf$", "", base_no_inf, flags=re.IGNORECASE)
    for candidate in (raw, base, base_no_inf, base_no_erf, base_no_inf_erf):
        key = normalize_key(candidate)
        if key:
            keys.add(key)
    return keys


def run_xlsx_import(
    db,
    filepath: str,
    row_progress_callback: Optional[Callable[[int, int], None]] = None,
) -> dict:
    """Führt den XLSX-Import durch und aktualisiert vorhandene Datensätze.

    Args:
        db: KarteikartenDB-Instanz
        filepath: Pfad zur XLSX-Datei
        row_progress_callback: optionale Funktion(current, total) für Fortschrittsanzeige

    Returns:
        dict mit 'updated', 'not_found', 'errors'

    Raises:
        ImportError: wenn openpyxl nicht installiert ist
        ValueError: wenn benötigte Spalten fehlen oder kein Header gefunden wird
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=filepath, read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if row and any(cell is not None for cell in row):
            header_row = row
            break

    if not header_row:
        raise ValueError("Keine Header-Zeile in der XLSX-Datei gefunden.")

    headers = {str(name).strip(): idx for idx, name in enumerate(header_row) if name is not None}

    required = [
        "Karteikarte", "Jahr", "Datum Taufe", "Datum Geburt", "Seite", "Nummer",
        "Karteikartentext", "Vorname Täufling", "Klarname", "Vorname Vater",
        "Geschlecht Täufling", "Kirchenbucheintrag",
    ]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError("Fehlende Spalten in der XLSX-Datei:\n" + "\n".join(missing))

    cursor = db.conn.cursor()
    cursor.execute(
        "SELECT id, dateiname FROM karteikarten WHERE dateiname IS NOT NULL AND dateiname <> ''"
    )
    key_to_ids: dict[str, list[int]] = {}
    for record_id, name in cursor.fetchall():
        for key in build_match_keys(name):
            key_to_ids.setdefault(key, []).append(record_id)

    updated = 0
    not_found = 0
    not_found_names: list[str] = []
    errors = 0
    total_rows = max(ws.max_row - 1, 0)
    created_at = "2026-01-16 00:00:00"

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if row_progress_callback:
            row_progress_callback(row_index, total_rows)
        try:
            dateiname = normalize_text(row[headers["Karteikarte"]])
            if not dateiname:
                continue

            row_keys = build_match_keys(dateiname)
            matched_ids = []
            for key in row_keys:
                if key in key_to_ids:
                    matched_ids.extend(key_to_ids[key])
            matched_ids = list(dict.fromkeys(matched_ids))

            if not matched_ids:
                not_found += 1
                not_found_names.append(dateiname)
                continue

            jahr = normalize_year(row[headers["Jahr"]])
            datum_taufe = normalize_date(row[headers["Datum Taufe"]], wb.epoch)
            datum_geburt = normalize_date(row[headers["Datum Geburt"]], wb.epoch)
            datum = datum_taufe or datum_geburt
            iso_datum = iso_from_datum(datum)
            seite = normalize_number(row[headers["Seite"]])
            nummer = normalize_number(row[headers["Nummer"]])
            erkannter_text = normalize_text(row[headers["Karteikartentext"]])
            vorname = normalize_text(row[headers["Vorname Täufling"]])
            nachname = normalize_text(row[headers["Klarname"]])
            partner = normalize_text(row[headers["Vorname Vater"]])
            stand = stand_from_gender(row[headers["Geschlecht Täufling"]])
            kirchenbuchtext = normalize_text(row[headers["Kirchenbucheintrag"]])
            mutter_vorname = (
                normalize_text(row[headers["Vorname Mutter"]])
                if "Vorname Mutter" in headers else None
            )

            for record_id in matched_ids:
                cursor.execute(
                    """
                    UPDATE karteikarten
                    SET kirchengemeinde = ?,
                        ereignis_typ = ?,
                        jahr = ?,
                        datum = ?,
                        seite = ?,
                        nummer = ?,
                        erkannter_text = ?,
                        ocr_methode = ?,
                        erstellt_am = ?,
                        aktualisiert_am = CURRENT_TIMESTAMP,
                        iso_datum = ?,
                        vorname = ?,
                        nachname = ?,
                        partner = ?,
                        todestag = ?,
                        ort = ?,
                        stand = ?,
                        kirchenbuchtext = ?,
                        geb_jahr_gesch = ?,
                        mutter_vorname = ?,
                        datum_geburt = ?,
                        version = COALESCE(version, 1) + 1,
                        sync_status = 'pending',
                        updated_by = 'erkennung'
                    WHERE id = ?
                    """,
                    (
                        "ev. Kb. Wetzlar",
                        "Taufe",
                        jahr,
                        datum,
                        seite,
                        nummer,
                        erkannter_text,
                        "Import",
                        created_at,
                        iso_datum,
                        vorname,
                        nachname,
                        partner,
                        to_ymd_dot(datum_taufe),
                        "Wetzlar",
                        stand,
                        kirchenbuchtext,
                        jahr,
                        mutter_vorname,
                        to_ymd_dot(datum_taufe or datum_geburt),
                        record_id,
                    ),
                )
            updated += len(matched_ids)
        except Exception:
            errors += 1

    db.conn.commit()

    all_record_ids = [rid for ids in key_to_ids.values() for rid in ids]
    for record_id in all_record_ids:
        try:
            db.mark_record_for_sync(record_id)
        except Exception:
            pass

    return {"updated": updated, "not_found": not_found, "not_found_names": not_found_names, "errors": errors}
